import os #for os related function
import sys #for commandline input
import re # for regular expression
import docx #for docx files
import subprocess #for linux commands
import random, string #to genrate random names
import openpyxl as excel # to write info in excel sheet
from threading import * # for multithreading
import datetime


class CollectInfo:
        #re to find phone no
	phone= re.compile('((?:\+?(?: |-|\.)?\d{1,2}(?: |-|\.)?)?(?:\(?:?\d{3}\)?|\d{3})(?: |-|\.)?(?:\d{3}(?: |-|\.)?\d{4}))')
	#re to find email
	email = re.compile("([a-z0-9]+[_a-z0-9\.-]*[a-z0-9]+@[a-z0-9-]+(?:\.[a-z0-9-]+)*(?:\.[a-z]{2,4}))")
	pdfFiles=[] # to store pdf filenames
	DocFiles=[] # to store docx filenames
	path1="" #path to folder where resume files are presesnt
	infoCollection=[] # list of collected info
	
	#constructor 
	def __init__(self,path1):
		self.path1 =path1
	
	#function to traverse all available files in folder
	def collectfiles(self):
		
		for (dirname,dirs,files) in os.walk(self.path1):
			for filename in files:
				if(filename.endswith('.docx')):
					x = os.path.join(dirname,filename)
					y=os.path.abspath(x)
					self.DocFiles.append(y)
					
				elif(filename.endswith('.pdf')):
					x=os.path.abspath(os.path.join(dirname,filename))
					self.pdfFiles.append(x)
						
		
	#to print pdf and docs files name
	def printDetails(self):
		fileObj = open("FilesVisited.txt",'w')
		str1 =datetime.datetime.now().strftime ("%Y-%m-%d %H:%M:%S")
		fileObj.write((str1+"\n\n"))
		fileObj.write("List of all pdf and docx files\n\n")
		i=1
		for files in self.pdfFiles:
			fileObj.write((str(i)+". "+files))
			fileObj.write("\n")
			i+=1			
		for files in self.DocFiles:
			fileObj.write((str(i)+". "+files))
			fileObj.write("\n")
			i+=1
		fileObj.close()
	#to change director and create directory
	def osStuff(self):
		pathHome = os.path.expanduser('~')+'/Desktop'
		os.chdir(pathHome)
		str1 = "TextFiles"
		Thread(target=self.printDetails()).start()
		try:
			os.mkdir(str1)
		except FileExistsError:
			str1= 'TextFiles'+''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
			os.mkdir(str1)
		os.chdir(str1) 
		return str1
	
	#function to create textfiles for pdf content	
	def createTextfiles(self,cmd):
		subprocess.call(cmd,shell=True)
	
	#function to find email name and phone from textfiles 
	def extractInfotxt(self,files):
		flg1=0
		info={}
		try:
			fileObj = open(files)
		except:
			print("file not found ")
			return
		for line in fileObj:
			m =re.findall(self.email,line)
			m1 = re.findall(self.phone,line)
			if m :  
				namelst=[]
				val = ','.join(map(str,m))
				i=0
				while(not (val[i]=="@" )):
					if(val[i].isalpha()):
						namelst.append(val[i])
					i+=1			
				#c1 = sheet.cell(row =cnt , column =4)
				#c1.value=val
				info['email']=val	
				#print(m)
				#print(line)
				valN=''.join(map(str,namelst))
				#sheet.cell(row =cnt , column =2,value=valN)
				info['name']=valN
				#sheet.cell(row =cnt , column =1,value=cnt)
				
				flg1+=1			
			if m1:
				val = ','.join(map(str,m1))
				#sheet.cell(row=cnt,column=3 ).value=val
				info['mobile']=val
				#fout.write(''.join(map(str,m1)))
				#print(m1)
				flg1+=1
			if (flg1>=2):
				break
		self.infoCollection.append(info)
	
	#function to collect required info from pdfFiles
	def getInfopdf(self):
		str1 = self.osStuff()
		i=1
		for files in self.pdfFiles:
			cmd = "pdftotext "+files+" "+"file"+str(i)+".txt"
			#Thread(target=self.createTextfiles,args=(cmd,)).start()
			try:
				subprocess.call(cmd,shell=True)
			except:
				print("some error in command")
			i+=1
		for files in os.listdir('.'):
			try:
				Thread(target=self.extractInfotxt,args=(files,)).start()
			except e:
				print("some thread exception ",e)
		#wb.save("demo.xlsx")
		#endtime= time.time()
		#print("total time take is " ,endtime-begintime)
	
	#function to collect required info from docFiles
	def findinfor_docfile(self,docfilename):
		try:
			resume = docx.Document(docfilename)
		except :
			print("doc files has some error")
			return
		flg1=0
		info={}
		for para in resume.paragraphs:
			#print(para.text)
			m =re.findall(self.email,para.text)
			m1 = re.findall(self.phone,para.text)
			if m :  
				namelst=[]
				val = ','.join(map(str,m))
				i=0
				while(not (val[i]=="@" )):
					if(val[i].isalpha()):
						namelst.append(val[i])
					i+=1			
				#c1 = sheet.cell(row =cnt , column =4)
				#c1.value=val
				info['email']=val	
				#print(m)
				#print(line)
				valN=''.join(map(str,namelst))
				#sheet.cell(row =cnt , column =2,value=valN)
				info['name']=valN
				#sheet.cell(row =cnt , column =1,value=cnt)	
				flg1+=1			
			if m1:
				val = ','.join(map(str,m1))
				#sheet.cell(row=cnt,column=3 ).value=val
				info['mobile']=val
				#fout.write(''.join(map(str,m1)))
				#print(m1)
				flg1+=1
			if (flg1>=2):
				break
		self.infoCollection.append(info)

	def getInfodocs(self):
		for filename in self.DocFiles:
			try:
				Thread(target=self.findinfor_docfile,args=(filename,)).start()
			except e:
				print(e)
	#function to create excel sheet
	def createExcelSheet(self):
		wb = excel.Workbook()
		sheet = wb.active
		sheet.title = "DataCollected"
		cnt=1
		for info in self.infoCollection:
			sheet.cell(row=cnt,column=1,value=cnt)
			try:
				sheet.cell(row=cnt,column=2,value=info['name'])
			except KeyError:
				sheet.cell(row=cnt,column=2,value="NULL")
			try:
				sheet.cell(row=cnt,column=3,value=info['email'])
			except KeyError:
				sheet.cell(row=cnt,column=3,value="NULL")
			try:
				sheet.cell(row=cnt,column=4,value=info['mobile'])
			except KeyError:
				sheet.cell(row=cnt,column=4,value="NULL")
			#print(cnt," ",info) # to print all collected info name ,email and phone
			cnt+=1
			
		wb.save("demo.xlsx")

#-----------------------------------------------------------------------------------
#---------------------------#
#main function
#---------------------------#

def main():
	try:
		path1 = sys.argv[1]
	except IndexError:
		print("enter path of directory where files are stored to retrieve information ")
		path1 = input()
	print("\n"*3)
	print("wait for sometime while we gather required information")	
	obj = CollectInfo(path1)
	obj.collectfiles()
	#obj.printDetails()
	t1=Thread(target=obj.getInfopdf)
	t2=Thread(target=obj.getInfodocs)
	t1.start()
	t2.start()
	t1.join()
	t2.join()
	while(active_count()>1):# wait to finish working of all threads
		continue	
	obj.createExcelSheet()
	
	print("#############################")
	print("Important message ")
	print("#############################")
	print("Please check your desktop directory")
	print("a directory with name TextFiles is created and and demo.xlx files with all data is there ")
	print("A text file \"FilesVisited.txt\" contains list of files used to get information")
	print("thank you !!! Good Byye")
	print("\n"*2)

#------------------------------------------------------------------------------------------------	
if __name__== "__main__":
	main()
